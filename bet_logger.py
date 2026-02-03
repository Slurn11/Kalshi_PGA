import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

LOG_PATH = Path(__file__).parent / "bet_log.xlsx"
CSV_PATH = Path(__file__).parent / "bet_log.csv"

HEADERS = [
    "Timestamp",
    "Tournament",
    "Player",
    "Market Type",
    "Ticker",
    "DG Probability",
    "Kalshi Implied Prob",
    "Edge %",
    "Decision",
    "Confidence",
    "Suggested Stake %",
    "Reasoning",
    "Position",
    "Score to Par",
    "Round",
    "Holes Completed",
    "Outcome",
    "Profit/Loss",
]


def _get_workbook() -> Workbook:
    if LOG_PATH.exists():
        return load_workbook(str(LOG_PATH))
    wb = Workbook()
    ws = wb.active
    ws.title = "Bet Log"
    ws.append(HEADERS)
    # Bold headers
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    wb.save(str(LOG_PATH))
    return wb


def log_recommendation(
    player_name: str,
    market_type: str,
    market_ticker: str,
    dg_prob: float,
    kalshi_implied_prob: float,
    edge_pct: float,
    decision: str,
    confidence: float,
    suggested_stake_pct: float,
    reasoning: str,
    leaderboard_context: Optional[dict] = None,
    tournament_name: str = "",
):
    """Log a recommendation (BET, PASS, or WATCH) to bet_log.xlsx."""
    wb = _get_workbook()
    ws = wb.active

    ctx = leaderboard_context or {}
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        tournament_name,
        player_name,
        market_type,
        market_ticker,
        round(dg_prob * 100, 1),
        round(kalshi_implied_prob * 100, 1),
        round(edge_pct, 1),
        decision,
        round(confidence * 100),
        round(suggested_stake_pct, 1),
        reasoning[:200],
        ctx.get("position", ""),
        ctx.get("score_to_par", ""),
        ctx.get("round_number", ""),
        ctx.get("holes_completed", ""),
        "",  # Outcome — filled manually
        "",  # Profit/Loss — filled manually
    ]

    ws.append(row)
    wb.save(str(LOG_PATH))

    # Also write to CSV for easy viewing
    write_header = not CSV_PATH.exists()
    with open(CSV_PATH, "a", newline="") as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(HEADERS)
        writer.writerow(row)

    logger.info(f"Logged {decision} for {player_name} {market_type} to bet_log")


def get_historical_stats(market_type: Optional[str] = None) -> dict:
    """Get win/loss stats from bet_log.xlsx.

    Returns:
        {
            "type_wins": int, "type_total": int, "type_winrate": float,
            "type_pnl": float,
            "all_wins": int, "all_total": int, "all_winrate": float,
            "all_pnl": float,
        }
    """
    stats = {
        "type_wins": 0, "type_total": 0, "type_winrate": 0.0, "type_pnl": 0.0,
        "all_wins": 0, "all_total": 0, "all_winrate": 0.0, "all_pnl": 0.0,
    }

    if not LOG_PATH.exists():
        return stats

    wb = load_workbook(str(LOG_PATH), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    for row in rows:
        if len(row) < 18:
            continue

        row_decision = row[8]   # Decision column
        row_type = row[3]       # Market Type column
        outcome = row[16]       # Outcome column
        pnl = row[17]           # Profit/Loss column

        # Only count rows where outcome is filled and decision was BET
        if not outcome or row_decision != "BET":
            continue

        outcome_str = str(outcome).strip().upper()
        if outcome_str not in ("WIN", "LOSS"):
            continue

        is_win = outcome_str == "WIN"
        pnl_val = float(pnl) if pnl else 0.0

        # All bets
        stats["all_total"] += 1
        if is_win:
            stats["all_wins"] += 1
        stats["all_pnl"] += pnl_val

        # Type-specific
        if market_type and str(row_type).strip().lower() == market_type.lower():
            stats["type_total"] += 1
            if is_win:
                stats["type_wins"] += 1
            stats["type_pnl"] += pnl_val

    if stats["all_total"] > 0:
        stats["all_winrate"] = stats["all_wins"] / stats["all_total"]
    if stats["type_total"] > 0:
        stats["type_winrate"] = stats["type_wins"] / stats["type_total"]

    return stats
